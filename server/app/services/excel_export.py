"""
Excel export service for task and analytics documents.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.models import TaskModel, UserModel
from app.services.tenant_filter import TenantFilter

logger = logging.getLogger(__name__)

STATUS_COLORS = {
    "NEW": "FCA5A5",
    "IN_PROGRESS": "FCD34D",
    "DONE": "86EFAC",
    "CANCELLED": "D1D5DB",
}

PRIORITY_COLORS = {
    "PLANNED": "BBF7D0",
    "1": "BBF7D0",
    "CURRENT": "BFDBFE",
    "2": "BFDBFE",
    "URGENT": "FED7AA",
    "3": "FED7AA",
    "EMERGENCY": "FECACA",
    "4": "FECACA",
}

STATUS_LABELS = {
    "NEW": "Новая",
    "IN_PROGRESS": "В работе",
    "DONE": "Выполнена",
    "CANCELLED": "Отменена",
}

PRIORITY_LABELS = {
    "PLANNED": "Плановая",
    "1": "Плановая",
    "CURRENT": "Текущая",
    "2": "Текущая",
    "URGENT": "Срочная",
    "3": "Срочная",
    "EMERGENCY": "Аварийная",
    "4": "Аварийная",
}

DETAIL_HEADERS = [
    ("№", 8),
    ("Номер заявки", 15),
    ("Название", 35),
    ("Статус", 14),
    ("Приоритет", 14),
    ("Адрес", 40),
    ("Исполнитель", 20),
    ("Заказчик", 20),
    ("Телефон", 16),
    ("Создана", 18),
    ("Дата план.", 18),
    ("Завершена", 18),
    ("Время выполн. (ч)", 18),
    ("Сумма", 12),
    ("Описание", 50),
]


def export_tasks_to_excel(
    db: Session,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    assignee_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    tenant_user: Optional[UserModel] = None,
    *,
    document_title: str = "Реестр заявок",
    period_label: Optional[str] = None,
    generated_by: Optional[str] = None,
    organization_name: Optional[str] = None,
    worker_name: Optional[str] = None,
    overview_metrics: Optional[dict[str, Any]] = None,
    sla_metrics: Optional[dict[str, Any]] = None,
) -> io.BytesIO:
    """
    Export tasks as a polished Excel document.
    """
    tasks = _load_tasks(
        db=db,
        status=status,
        priority=priority,
        assignee_id=assignee_id,
        date_from=date_from,
        date_to=date_to,
        tenant_user=tenant_user,
    )

    wb = Workbook()
    styles = _build_styles()

    ws_cover = wb.active
    ws_cover.title = "Отчет"
    _add_cover_sheet(
        ws=ws_cover,
        styles=styles,
        document_title=document_title,
        period_label=period_label or "Без ограничения периода",
        generated_by=generated_by or "Система FieldWorker",
        organization_name=organization_name or "Все организации",
        worker_name=worker_name or "Все исполнители",
        tasks=tasks,
        overview_metrics=overview_metrics,
        sla_metrics=sla_metrics,
    )

    ws_summary = wb.create_sheet("Сводка")
    _add_summary_sheet(
        ws=ws_summary,
        tasks=tasks,
        styles=styles,
        overview_metrics=overview_metrics,
        sla_metrics=sla_metrics,
    )

    ws_details = wb.create_sheet("Заявки")
    _add_tasks_sheet(ws=ws_details, tasks=tasks, styles=styles)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("Excel export: %d tasks exported", len(tasks))
    return output


def _load_tasks(
    db: Session,
    *,
    status: Optional[str],
    priority: Optional[str],
    assignee_id: Optional[int],
    date_from: Optional[str],
    date_to: Optional[str],
    tenant_user: Optional[UserModel],
) -> list[TaskModel]:
    tenant = TenantFilter(tenant_user) if tenant_user is not None else None
    query = db.query(TaskModel).options(joinedload(TaskModel.assigned_user))
    if tenant is not None:
        query = tenant.apply(query, TaskModel)

    if status:
        query = query.filter(TaskModel.status == status)

    if priority:
        query = query.filter(TaskModel.priority.in_([priority, priority.upper()]))

    if assignee_id is not None:
        query = query.filter(TaskModel.assigned_user_id == assignee_id)

    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from)
            query = query.filter(TaskModel.created_at >= dt_from)
        except ValueError:
            pass

    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to)
            query = query.filter(TaskModel.created_at <= dt_to)
        except ValueError:
            pass

    return query.order_by(TaskModel.created_at.desc()).all()


def _build_styles() -> dict[str, Any]:
    border = Border(
        left=Side(style="thin", color="D7DCE3"),
        right=Side(style="thin", color="D7DCE3"),
        top=Side(style="thin", color="D7DCE3"),
        bottom=Side(style="thin", color="D7DCE3"),
    )
    return {
        "title_font": Font(name="Calibri", size=20, bold=True, color="FFFFFF"),
        "section_font": Font(name="Calibri", size=12, bold=True, color="0F172A"),
        "header_font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        "label_font": Font(name="Calibri", size=10, bold=True, color="334155"),
        "value_font": Font(name="Calibri", size=10, color="0F172A"),
        "body_font": Font(name="Calibri", size=10, color="0F172A"),
        "muted_font": Font(name="Calibri", size=10, color="64748B"),
        "header_fill": PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid"),
        "title_fill": PatternFill(start_color="0F172A", end_color="1E293B", fill_type="solid"),
        "soft_fill": PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"),
        "accent_fill": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "zebra_fill": PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"),
        "border": border,
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "right": Alignment(horizontal="right", vertical="center", wrap_text=True),
    }


def _add_cover_sheet(
    ws,
    *,
    styles: dict[str, Any],
    document_title: str,
    period_label: str,
    generated_by: str,
    organization_name: str,
    worker_name: str,
    tasks: list[TaskModel],
    overview_metrics: Optional[dict[str, Any]],
    sla_metrics: Optional[dict[str, Any]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = document_title
    title_cell.font = styles["title_font"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = styles["center"]

    ws.merge_cells("A3:H3")
    subtitle = ws["A3"]
    subtitle.value = "Профессиональный аналитический документ FieldWorker"
    subtitle.font = styles["muted_font"]
    subtitle.alignment = styles["center"]

    for column, width in {
        "A": 18,
        "B": 22,
        "C": 18,
        "D": 22,
        "E": 18,
        "F": 22,
        "G": 18,
        "H": 22,
    }.items():
        ws.column_dimensions[column].width = width

    metadata_rows = [
        ("Период", period_label),
        ("Исполнитель", worker_name),
        ("Организация", organization_name),
        ("Сформирован", datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")),
        ("Подготовил", generated_by),
        ("Всего заявок", str(len(tasks))),
    ]

    start_row = 5
    for index, (label, value) in enumerate(metadata_rows):
        row = start_row + index
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        label_cell = ws.cell(row=row, column=2, value=label)
        value_cell = ws.cell(row=row, column=4, value=value)
        label_cell.font = styles["label_font"]
        label_cell.fill = styles["soft_fill"]
        label_cell.alignment = styles["left"]
        value_cell.font = styles["value_font"]
        value_cell.alignment = styles["left"]
        for col in range(2, 8):
            ws.cell(row=row, column=col).border = styles["border"]

    overview = overview_metrics or {}
    reports_summary = overview.get("summary", {})
    completion_time = overview.get("completion_time") or {}
    sla_overview = (sla_metrics or {}).get("overview", {})
    sla_timing = (sla_metrics or {}).get("timing", {})

    row = 13
    row = _write_metric_table(
        ws=ws,
        styles=styles,
        start_row=row,
        title="Операционный обзор",
        rows=[
            ("Всего заявок", reports_summary.get("total_tasks", len(tasks))),
            ("Выполнено", reports_summary.get("completed_tasks", 0)),
            ("Среднее в день", reports_summary.get("avg_tasks_per_day", 0)),
            ("Среднее время, ч", completion_time.get("avg_hours", 0)),
        ],
    )
    row += 1
    _write_metric_table(
        ws=ws,
        styles=styles,
        start_row=row,
        title="SLA и сроки",
        rows=[
            ("SLA compliance", f"{sla_overview.get('sla_compliance_rate', 0)}%"),
            ("Просроченные активные", sla_overview.get("active_overdue", 0)),
            ("Просроченные завершенные", sla_overview.get("overdue_tasks", 0)),
            ("Среднее время SLA, ч", sla_timing.get("avg_completion_hours", 0)),
        ],
    )


def _add_summary_sheet(
    ws,
    *,
    tasks: list[TaskModel],
    styles: dict[str, Any],
    overview_metrics: Optional[dict[str, Any]],
    sla_metrics: Optional[dict[str, Any]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "Сводка по заявкам и SLA"
    title.font = Font(name="Calibri", size=16, bold=True, color="0F172A")
    title.alignment = styles["left"]

    ws["A2"] = f"Дата формирования: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M UTC')}"
    ws["A2"].font = styles["muted_font"]

    row = 4
    row = _write_status_table(ws=ws, styles=styles, start_row=row, tasks=tasks)
    row += 2
    row = _write_priority_table(ws=ws, styles=styles, start_row=row, tasks=tasks)

    overview = overview_metrics or {}
    reports_summary = overview.get("summary", {})
    reports_workers = overview.get("by_worker") or []
    sla = sla_metrics or {}
    sla_overview = sla.get("overview", {})
    sla_timing = sla.get("timing", {})
    sla_priorities = sla.get("by_priority") or []

    _write_metric_table(
        ws=ws,
        styles=styles,
        start_row=4,
        title="Ключевые метрики",
        rows=[
            ("Всего заявок", reports_summary.get("total_tasks", len(tasks))),
            ("Выполнено", reports_summary.get("completed_tasks", 0)),
            ("% выполнения", f"{reports_summary.get('completion_rate', 0)}%"),
            ("SLA compliance", f"{sla_overview.get('sla_compliance_rate', 0)}%"),
            ("Активные просрочки", sla_overview.get("active_overdue", 0)),
            ("Медиана SLA, ч", sla_timing.get("median_completion_hours", 0)),
        ],
        start_column=5,
        end_column=7,
    )

    worker_start = 18
    ws.merge_cells(start_row=worker_start, start_column=1, end_row=worker_start, end_column=3)
    worker_title = ws.cell(row=worker_start, column=1, value="Исполнители")
    worker_title.font = styles["section_font"]
    worker_title.alignment = styles["left"]

    for offset, header in enumerate(["Исполнитель", "Всего", "% выполнения"], start=1):
        cell = ws.cell(row=worker_start + 1, column=offset, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    for index, worker in enumerate(reports_workers[:8], start=worker_start + 2):
        completion_rate = 0
        if worker.get("total", 0):
            completion_rate = round(worker.get("completed", 0) / worker.get("total", 1) * 100, 1)
        values = [
            worker.get("user_name", "—"),
            worker.get("total", 0),
            f"{completion_rate}%",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=index, column=col, value=value)
            cell.font = styles["body_font"]
            cell.alignment = styles["left"] if col == 1 else styles["center"]
            cell.border = styles["border"]
            if index % 2 == 0:
                cell.fill = styles["zebra_fill"]

    sla_priority_start = 18
    ws.merge_cells(start_row=sla_priority_start, start_column=5, end_row=sla_priority_start, end_column=7)
    sla_title = ws.cell(row=sla_priority_start, column=5, value="SLA по приоритетам")
    sla_title.font = styles["section_font"]
    sla_title.alignment = styles["left"]

    for offset, header in enumerate(["Приоритет", "SLA %", "Норма"], start=5):
        cell = ws.cell(row=sla_priority_start + 1, column=offset, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    for index, priority_row in enumerate(sla_priorities[:4], start=sla_priority_start + 2):
        values = [
            priority_row.get("label", priority_row.get("priority", "—")),
            f"{priority_row.get('sla_compliance_rate', 0)}%",
            priority_row.get("sla_hours", 0),
        ]
        for col, value in enumerate(values, start=5):
            cell = ws.cell(row=index, column=col, value=value)
            cell.font = styles["body_font"]
            cell.alignment = styles["left"] if col == 5 else styles["center"]
            cell.border = styles["border"]
            if index % 2 == 0:
                cell.fill = styles["zebra_fill"]


def _add_tasks_sheet(ws, *, tasks: list[TaskModel], styles: dict[str, Any]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_HEADERS))}{max(len(tasks) + 1, 2)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    for col_idx, (header_text, width) in enumerate(DETAIL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, task in enumerate(tasks, start=2):
        assignee_name = ""
        if task.assigned_user:
            assignee_name = task.assigned_user.full_name or task.assigned_user.username

        planned_str = ""
        if task.planned_date:
            if isinstance(task.planned_date, datetime):
                planned_str = task.planned_date.strftime("%d.%m.%Y")
            else:
                planned_str = str(task.planned_date)

        completion_hours: float | str = ""
        if task.created_at and task.completed_at:
            completion_hours = round((task.completed_at - task.created_at).total_seconds() / 3600, 1)

        values = [
            row_idx - 1,
            task.task_number or "",
            task.title or "",
            STATUS_LABELS.get(task.status, task.status),
            PRIORITY_LABELS.get(str(task.priority).upper(), str(task.priority or "")),
            task.raw_address or "",
            assignee_name,
            task.customer_name or "",
            task.customer_phone or "",
            task.created_at.strftime("%d.%m.%Y %H:%M") if task.created_at else "",
            planned_str,
            task.completed_at.strftime("%d.%m.%Y %H:%M") if task.completed_at else "",
            completion_hours,
            getattr(task, "payment_amount", None) or "",
            task.description or "",
        ]

        zebra = row_idx % 2 == 0
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = styles["body_font"]
            cell.alignment = styles["left"] if col_idx in {2, 3, 6, 7, 8, 9, 15} else styles["center"]
            cell.border = styles["border"]
            if zebra:
                cell.fill = styles["zebra_fill"]

        status_fill = STATUS_COLORS.get(task.status)
        if status_fill:
            ws.cell(row=row_idx, column=4).fill = PatternFill(
                start_color=status_fill,
                end_color=status_fill,
                fill_type="solid",
            )
            ws.cell(row=row_idx, column=4).font = Font(name="Calibri", size=10, bold=True, color="0F172A")

        priority_key = str(task.priority).upper() if task.priority else ""
        priority_fill = PRIORITY_COLORS.get(priority_key)
        if priority_fill:
            ws.cell(row=row_idx, column=5).fill = PatternFill(
                start_color=priority_fill,
                end_color=priority_fill,
                fill_type="solid",
            )
            ws.cell(row=row_idx, column=5).font = Font(name="Calibri", size=10, bold=True, color="0F172A")


def _write_metric_table(
    ws,
    *,
    styles: dict[str, Any],
    start_row: int,
    title: str,
    rows: list[tuple[str, Any]],
    start_column: int = 2,
    end_column: int = 4,
) -> int:
    ws.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=start_row,
        end_column=end_column,
    )
    title_cell = ws.cell(row=start_row, column=start_column, value=title)
    title_cell.font = styles["section_font"]
    title_cell.alignment = styles["left"]

    current_row = start_row + 1
    for label, value in rows:
        label_cell = ws.cell(row=current_row, column=start_column, value=label)
        value_cell = ws.cell(row=current_row, column=start_column + 1, value=value)
        label_cell.font = styles["label_font"]
        label_cell.fill = styles["soft_fill"]
        label_cell.alignment = styles["left"]
        value_cell.font = styles["value_font"]
        value_cell.alignment = styles["left"]
        for col in range(start_column, end_column + 1):
            ws.cell(row=current_row, column=col).border = styles["border"]
            if col > start_column + 1:
                ws.cell(row=current_row, column=col).fill = styles["accent_fill"]
        current_row += 1

    return current_row


def _write_status_table(ws, *, styles: dict[str, Any], start_row: int, tasks: list[TaskModel]) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    title = ws.cell(row=start_row, column=1, value="По статусам")
    title.font = styles["section_font"]
    title.alignment = styles["left"]

    for col, header in enumerate(["Статус", "Количество", "%"], start=1):
        cell = ws.cell(row=start_row + 1, column=col, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    row = start_row + 2
    total = len(tasks)
    for status_key, status_label in STATUS_LABELS.items():
        count = sum(1 for task in tasks if task.status == status_key)
        percent = round(count / total * 100, 1) if total else 0
        values = [status_label, count, f"{percent}%"]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = styles["body_font"]
            cell.alignment = styles["left"] if col == 1 else styles["center"]
            cell.border = styles["border"]
            if row % 2 == 0:
                cell.fill = styles["zebra_fill"]
        row += 1
    return row


def _write_priority_table(ws, *, styles: dict[str, Any], start_row: int, tasks: list[TaskModel]) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    title = ws.cell(row=start_row, column=1, value="По приоритетам")
    title.font = styles["section_font"]
    title.alignment = styles["left"]

    for col, header in enumerate(["Приоритет", "Количество", "%"], start=1):
        cell = ws.cell(row=start_row + 1, column=col, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    row = start_row + 2
    total = len(tasks)
    priority_order = ["PLANNED", "CURRENT", "URGENT", "EMERGENCY"]
    rank_map = {"PLANNED": "1", "CURRENT": "2", "URGENT": "3", "EMERGENCY": "4"}
    for priority_key in priority_order:
        count = sum(
            1 for task in tasks if str(task.priority).upper() in {priority_key, rank_map[priority_key]}
        )
        percent = round(count / total * 100, 1) if total else 0
        values = [PRIORITY_LABELS[priority_key], count, f"{percent}%"]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = styles["body_font"]
            cell.alignment = styles["left"] if col == 1 else styles["center"]
            cell.border = styles["border"]
            if row % 2 == 0:
                cell.fill = styles["zebra_fill"]
        row += 1
    return row
