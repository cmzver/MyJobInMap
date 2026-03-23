"""
Tests for unified analytics API.
"""

import io
from datetime import datetime, timedelta

from openpyxl import load_workbook

from app.models import TaskModel, UserModel, UserRole
from app.services.auth import get_password_hash


def test_analytics_requires_auth(client):
    response = client.get("/api/analytics?period=month")
    assert response.status_code == 401


def test_analytics_combines_reports_and_sla(client_with_auth, sample_tasks_for_reports):
    response = client_with_auth.get("/api/analytics?period=all")

    assert response.status_code == 200
    data = response.json()

    assert "reports" in data
    assert "sla" in data
    assert (
        data["reports"]["summary"]["total_tasks"]
        == data["sla"]["overview"]["total_tasks"]
    )
    assert (
        data["reports"]["summary"]["completed_tasks"]
        == data["sla"]["overview"]["completed_tasks"]
    )


def test_analytics_week_uses_shared_period_days(
    client_with_auth, sample_tasks_for_reports
):
    response = client_with_auth.get("/api/analytics?period=week")

    assert response.status_code == 200
    data = response.json()

    assert data["reports"]["summary"]["period_days"] == data["sla"]["period"]["days"]
    assert 1 <= data["sla"]["period"]["days"] <= 7


def test_sla_supports_extended_periods(client_with_auth, sample_tasks_for_reports):
    for period in ("yesterday", "all"):
        response = client_with_auth.get(f"/api/sla?period={period}")
        assert response.status_code == 200


def test_analytics_export_endpoints(client_with_auth):
    document_response = client_with_auth.get("/api/analytics/export?period=month")
    excel_response = client_with_auth.get("/api/analytics/export/excel?period=month")

    assert document_response.status_code == 200
    assert "spreadsheetml" in document_response.headers.get("content-type", "")
    workbook = load_workbook(io.BytesIO(document_response.content))
    assert {"Отчет", "Сводка", "Заявки"}.issubset(set(workbook.sheetnames))

    assert excel_response.status_code == 200
    assert "spreadsheetml" in excel_response.headers.get("content-type", "")


def test_analytics_export_rejects_unsupported_format(client_with_auth):
    response = client_with_auth.get("/api/analytics/export?period=month&format=pdf")

    assert response.status_code == 400
    assert "xlsx" in response.json()["detail"]


def test_sla_worker_filter_applies_to_trends(client_with_auth, db_session):
    now = datetime.now()
    worker_one = UserModel(
        username="worker-one",
        password_hash=get_password_hash("worker-one"),
        full_name="Worker One",
        role=UserRole.WORKER.value,
        is_active=True,
    )
    worker_two = UserModel(
        username="worker-two",
        password_hash=get_password_hash("worker-two"),
        full_name="Worker Two",
        role=UserRole.WORKER.value,
        is_active=True,
    )
    db_session.add_all([worker_one, worker_two])
    db_session.commit()
    db_session.refresh(worker_one)
    db_session.refresh(worker_two)

    db_session.add_all(
        [
            TaskModel(
                title="Worker one task",
                description="Test",
                raw_address="Address 1",
                status="DONE",
                priority="CURRENT",
                created_at=now - timedelta(days=2),
                completed_at=now - timedelta(days=1),
                updated_at=now - timedelta(days=1),
                assigned_user_id=worker_one.id,
            ),
            TaskModel(
                title="Worker two task",
                description="Test",
                raw_address="Address 2",
                status="DONE",
                priority="CURRENT",
                created_at=now - timedelta(days=2),
                completed_at=now - timedelta(hours=12),
                updated_at=now - timedelta(hours=12),
                assigned_user_id=worker_two.id,
            ),
        ]
    )
    db_session.commit()

    response = client_with_auth.get(f"/api/sla?period=all&worker_id={worker_one.id}")

    assert response.status_code == 200
    data = response.json()
    assert data["overview"]["total_tasks"] == 1
    assert sum(item["created"] for item in data["trends"]) == 1
    assert sum(item["completed"] for item in data["trends"]) == 1
