"""
Tests for SLA API, Excel Export, WebSocket, and Prometheus Metrics
==================================================================
"""

import asyncio
import io
import json
from datetime import datetime, timedelta, timezone
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi.testclient import TestClient

from app.models import OrganizationModel, UserModel, UserRole
from app.models.task import TaskModel
from app.services.auth import get_password_hash
from app.services.excel_export import export_tasks_to_excel
from app.services.sla_service import _get_sla_hours, get_sla_metrics
from app.services.websocket_manager import (
    ConnectionManager, _event, broadcast_chat_conversation_updated)


def run_async(coro):
    """Запуск async функции в sync тесте."""
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


# ============================================================================
# SLA Service — Unit Tests
# ============================================================================


class TestSlaHours:
    """Тесты для SLA нормативов по приоритету."""

    def test_planned_sla(self):
        assert _get_sla_hours("PLANNED") == 168

    def test_current_sla(self):
        assert _get_sla_hours("CURRENT") == 48

    def test_urgent_sla(self):
        assert _get_sla_hours("URGENT") == 8

    def test_emergency_sla(self):
        assert _get_sla_hours("EMERGENCY") == 4

    def test_numeric_sla(self):
        assert _get_sla_hours("1") == 168
        assert _get_sla_hours("2") == 48
        assert _get_sla_hours("3") == 8
        assert _get_sla_hours("4") == 4

    def test_unknown_priority_default(self):
        assert _get_sla_hours("UNKNOWN") == 48

    def test_none_priority_default(self):
        assert _get_sla_hours(None) == 48


class TestSlaApi:
    """Тесты SLA API эндпоинтов."""

    def test_sla_endpoint_unauthenticated(self, client):
        """SLA API требует аутентификации."""
        response = client.get("/api/sla")
        assert response.status_code == 401

    def test_sla_endpoint_worker_forbidden(self, client_with_worker):
        """Worker не имеет доступа к SLA."""
        response = client_with_worker.get("/api/sla")
        assert response.status_code == 403

    def test_sla_endpoint_admin_ok(self, client_with_auth):
        """Admin имеет доступ к SLA."""
        response = client_with_auth.get("/api/sla?period=month")
        assert response.status_code == 200
        data = response.json()
        assert "overview" in data
        assert "timing" in data
        assert "by_priority" in data
        assert "trends" in data

    def test_sla_endpoint_dispatcher_ok(self, client_with_dispatcher):
        """Dispatcher имеет доступ к SLA."""
        response = client_with_dispatcher.get("/api/sla?period=week")
        assert response.status_code == 200

    def test_sla_overview_structure(self, client_with_auth):
        """Проверка структуры overview."""
        response = client_with_auth.get("/api/sla?period=month")
        data = response.json()
        overview = data["overview"]
        assert "total_tasks" in overview
        assert "completed_tasks" in overview
        assert "sla_compliance_rate" in overview
        assert "overdue_tasks" in overview
        assert "active_overdue" in overview

    def test_sla_with_tasks(self, client_with_auth, sample_tasks_for_reports):
        """SLA с данными."""
        response = client_with_auth.get("/api/sla?period=month")
        data = response.json()
        assert data["overview"]["total_tasks"] >= 4

    def test_sla_by_priority(self, client_with_auth, sample_tasks_for_reports):
        """SLA по приоритетам содержит 4 записи."""
        response = client_with_auth.get("/api/sla?period=month")
        data = response.json()
        assert len(data["by_priority"]) == 4
        priorities = [p["priority"] for p in data["by_priority"]]
        assert "PLANNED" in priorities
        assert "CURRENT" in priorities
        assert "URGENT" in priorities
        assert "EMERGENCY" in priorities

    def test_sla_periods(self, client_with_auth):
        """Разные периоды работают."""
        for period in ["today", "week", "month", "quarter", "year"]:
            response = client_with_auth.get(f"/api/sla?period={period}")
            assert response.status_code == 200

    def test_sla_invalid_period(self, client_with_auth):
        """Невалидный период → 422."""
        response = client_with_auth.get("/api/sla?period=invalid")
        assert response.status_code == 422

    def test_sla_timing_structure(self, client_with_auth):
        """Проверка структуры timing."""
        response = client_with_auth.get("/api/sla?period=month")
        data = response.json()
        timing = data["timing"]
        assert "avg_completion_hours" in timing
        assert "min_completion_hours" in timing
        assert "max_completion_hours" in timing
        assert "median_completion_hours" in timing

    def test_sla_is_tenant_scoped(self, client, db_session):
        org1 = OrganizationModel(name="SLA Org 1", slug="sla-org-1", is_active=True)
        org2 = OrganizationModel(name="SLA Org 2", slug="sla-org-2", is_active=True)
        admin = UserModel(
            username="sla_admin",
            password_hash=get_password_hash("secret123"),
            full_name="SLA Admin",
            role=UserRole.ADMIN.value,
            is_active=True,
            organization=org1,
        )
        worker1 = UserModel(
            username="sla_worker_1",
            password_hash=get_password_hash("secret123"),
            full_name="SLA Worker 1",
            role=UserRole.WORKER.value,
            is_active=True,
            organization=org1,
        )
        worker2 = UserModel(
            username="sla_worker_2",
            password_hash=get_password_hash("secret123"),
            full_name="SLA Worker 2",
            role=UserRole.WORKER.value,
            is_active=True,
            organization=org2,
        )
        now = datetime.now(timezone.utc)
        own_task = TaskModel(
            title="Own SLA Task",
            raw_address="Own Addr",
            status="DONE",
            priority="CURRENT",
            organization=org1,
            assigned_user=worker1,
            created_at=now - timedelta(days=2),
            completed_at=now - timedelta(days=1),
            updated_at=now - timedelta(days=1),
        )
        foreign_task = TaskModel(
            title="Foreign SLA Task",
            raw_address="Foreign Addr",
            status="DONE",
            priority="CURRENT",
            organization=org2,
            assigned_user=worker2,
            created_at=now - timedelta(days=2),
            completed_at=now - timedelta(days=1),
            updated_at=now - timedelta(days=1),
        )
        db_session.add_all(
            [org1, org2, admin, worker1, worker2, own_task, foreign_task]
        )
        db_session.commit()

        login = client.post(
            "/api/auth/login", data={"username": "sla_admin", "password": "secret123"}
        )
        assert login.status_code == 200
        headers = {"Authorization": f"Bearer {login.json()['access_token']}"}

        response = client.get("/api/sla?period=month", headers=headers)

        assert response.status_code == 200
        data = response.json()
        assert data["overview"]["total_tasks"] == 1
        assert data["overview"]["completed_tasks"] == 1
        assert len(data["by_worker"]) == 1
        assert data["by_worker"][0]["user_name"] == "SLA Worker 1"


class TestSlaService:
    """Тесты SLA сервиса — unit level."""

    def test_sla_metrics_empty_db(self, db_session):
        """SLA метрики для пустой БД."""
        result = get_sla_metrics(db_session, period="month")
        assert result["overview"]["total_tasks"] == 0
        assert result["overview"]["sla_compliance_rate"] == 0
        assert result["timing"]["avg_completion_hours"] == 0

    def test_sla_metrics_with_data(self, db_session, worker_user):
        """SLA метрики с данными."""
        now = datetime.now(timezone.utc)

        # Задача в SLA (PLANNED, выполнена за 24ч, SLA=168ч)
        task1 = TaskModel(
            title="Task In SLA",
            raw_address="Addr",
            status="DONE",
            priority="PLANNED",
            created_at=now - timedelta(hours=24),
            completed_at=now,
            assigned_user_id=worker_user.id,
        )

        # Задача вне SLA (EMERGENCY, выполнена за 8ч, SLA=4ч)
        task2 = TaskModel(
            title="Task Out SLA",
            raw_address="Addr",
            status="DONE",
            priority="EMERGENCY",
            created_at=now - timedelta(hours=8),
            completed_at=now,
            assigned_user_id=worker_user.id,
        )

        db_session.add_all([task1, task2])
        db_session.commit()

        result = get_sla_metrics(db_session, period="month")
        assert result["overview"]["completed_tasks"] == 2
        assert result["overview"]["sla_compliance_rate"] == 50.0  # 1 из 2 в SLA


# ============================================================================
# Excel Export — Tests
# ============================================================================


class TestExcelExport:
    """Тесты экспорта в Excel."""

    def test_export_empty(self, db_session):
        """Экспорт пустой БД — создаёт файл с заголовками."""
        output = export_tasks_to_excel(db_session)
        assert isinstance(output, io.BytesIO)
        assert output.getvalue()[:4] == b"PK\x03\x04"  # xlsx = zip

    def test_export_with_tasks(self, db_session, sample_tasks_for_reports):
        """Экспорт с данными."""
        output = export_tasks_to_excel(db_session)
        content = output.getvalue()
        assert len(content) > 100

    def test_export_filter_status(self, db_session, sample_tasks_for_reports):
        """Экспорт с фильтром по статусу."""
        output_all = export_tasks_to_excel(db_session)
        output_done = export_tasks_to_excel(db_session, status="DONE")
        # Файл с фильтром должен быть не больше (может быть меньше)
        assert len(output_done.getvalue()) <= len(output_all.getvalue()) + 100

    def test_export_api_unauthenticated(self, client):
        """Excel API требует аутентификации."""
        response = client.get("/api/reports/export/excel")
        assert response.status_code == 401

    def test_export_api_admin_ok(self, client_with_auth):
        """Admin может экспортировать Excel."""
        response = client_with_auth.get("/api/reports/export/excel?period=month")
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers.get("content-type", "")

    def test_export_api_filename(self, client_with_auth):
        """Проверка имени файла в Content-Disposition."""
        response = client_with_auth.get("/api/reports/export/excel?period=month")
        disposition = response.headers.get("content-disposition", "")
        assert "tasks_month_" in disposition
        assert ".xlsx" in disposition

    def test_export_xlsx_valid(self, db_session, sample_tasks_for_reports):
        """Экспортированный файл — валидный xlsx."""
        from openpyxl import load_workbook

        output = export_tasks_to_excel(db_session)
        wb = load_workbook(output)

        # Документ содержит титульный лист, сводку и реестр заявок
        assert len(wb.sheetnames) == 3
        assert "Отчет" in wb.sheetnames
        assert "Заявки" in wb.sheetnames
        assert "Сводка" in wb.sheetnames
        assert wb["Отчет"].cell(1, 1).value is not None

        ws = wb["Заявки"]
        # Заголовки в первой строке
        assert ws.cell(1, 1).value == "№"
        assert ws.cell(1, 2).value == "Номер заявки"
        assert ws.cell(1, 3).value == "Название"

        # Данные начинаются со 2-й строки
        assert ws.cell(2, 3).value is not None  # Название первой задачи


# ============================================================================
# WebSocket Manager — Unit Tests
# ============================================================================


class TestWebSocketManager:
    """Тесты WebSocket менеджера."""

    @pytest.fixture
    def manager(self):
        return ConnectionManager()

    def test_initial_state(self, manager):
        """Начальное состояние менеджера."""
        assert manager.active_connections_count == 0
        assert manager.active_users_count == 0

    def test_get_status(self, manager):
        """Статус менеджера."""
        status = manager.get_status()
        assert status["active_connections"] == 0
        assert status["unique_users"] == 0
        assert status["connected_user_ids"] == []

    def test_connect_disconnect(self, manager):
        """Подключение и отключение."""
        ws = AsyncMock()
        ws.accept = AsyncMock()

        run_async(manager.connect(ws, user_id=1))
        assert manager.active_connections_count == 1
        assert manager.active_users_count == 1

        run_async(manager.disconnect(ws))
        assert manager.active_connections_count == 0
        assert manager.active_users_count == 0

    def test_multiple_connections_same_user(self, manager):
        """Несколько соединений одного пользователя."""
        ws1 = AsyncMock()
        ws1.accept = AsyncMock()
        ws2 = AsyncMock()
        ws2.accept = AsyncMock()

        run_async(manager.connect(ws1, user_id=1))
        run_async(manager.connect(ws2, user_id=1))

        assert manager.active_connections_count == 2
        assert manager.active_users_count == 1

    def test_broadcast(self, manager):
        """Broadcast отправляет всем."""
        ws1 = AsyncMock()
        ws1.accept = AsyncMock()
        ws1.send_json = AsyncMock()
        ws2 = AsyncMock()
        ws2.accept = AsyncMock()
        ws2.send_json = AsyncMock()

        run_async(manager.connect(ws1, user_id=1))
        run_async(manager.connect(ws2, user_id=2))

        msg = {"type": "test", "data": {}}
        sent = run_async(manager.broadcast(msg))

        assert sent == 2
        ws1.send_json.assert_called_once_with(msg)
        ws2.send_json.assert_called_once_with(msg)

    def test_broadcast_exclude_user(self, manager):
        """Broadcast с исключением пользователя."""
        ws1 = AsyncMock()
        ws1.accept = AsyncMock()
        ws1.send_json = AsyncMock()
        ws2 = AsyncMock()
        ws2.accept = AsyncMock()
        ws2.send_json = AsyncMock()

        run_async(manager.connect(ws1, user_id=1))
        run_async(manager.connect(ws2, user_id=2))

        msg = {"type": "test", "data": {}}
        sent = run_async(manager.broadcast(msg, exclude_user_id=1))

        assert sent == 1
        ws1.send_json.assert_not_called()
        ws2.send_json.assert_called_once_with(msg)

    def test_broadcast_scoped_by_organization(self, manager):
        """Broadcast не должен уходить в другую организацию."""
        ws1 = AsyncMock()
        ws1.accept = AsyncMock()
        ws1.send_json = AsyncMock()
        ws2 = AsyncMock()
        ws2.accept = AsyncMock()
        ws2.send_json = AsyncMock()
        ws3 = AsyncMock()
        ws3.accept = AsyncMock()
        ws3.send_json = AsyncMock()

        run_async(manager.connect(ws1, user_id=1, organization_id=10))
        run_async(manager.connect(ws2, user_id=2, organization_id=20))
        run_async(
            manager.connect(ws3, user_id=3, organization_id=None, is_superadmin=True)
        )

        msg = {"type": "test", "data": {}}
        sent = run_async(manager.broadcast(msg, organization_id=10))

        assert sent == 2
        ws1.send_json.assert_called_once_with(msg)
        ws2.send_json.assert_not_called()
        ws3.send_json.assert_called_once_with(msg)

    def test_send_to_user(self, manager):
        """Отправка конкретному пользователю."""
        ws = AsyncMock()
        ws.accept = AsyncMock()
        ws.send_json = AsyncMock()

        run_async(manager.connect(ws, user_id=42))

        msg = {"type": "test"}
        sent = run_async(manager.send_to_user(42, msg))

        assert sent == 1
        ws.send_json.assert_called_once_with(msg)

    def test_send_to_nonexistent_user(self, manager):
        """Отправка несуществующему пользователю."""
        sent = run_async(manager.send_to_user(999, {"type": "test"}))
        assert sent == 0

    def test_stale_connection_cleanup(self, manager):
        """Очистка мёртвых соединений при broadcast."""
        ws = AsyncMock()
        ws.accept = AsyncMock()
        ws.send_json = AsyncMock(side_effect=Exception("Connection closed"))

        run_async(manager.connect(ws, user_id=1))
        assert manager.active_connections_count == 1

        run_async(manager.broadcast({"type": "test"}))

        # Мёртвое соединение удалено
        assert manager.active_connections_count == 0


class TestWsEvent:
    """Тесты формата WebSocket событий."""

    def test_event_structure(self):
        """Стандартная структура события."""
        evt = _event("task_created", {"task_id": 1})
        assert evt["type"] == "task_created"
        assert evt["data"]["task_id"] == 1
        assert "timestamp" in evt

    def test_event_timestamp_iso(self):
        """Timestamp в ISO формате."""
        evt = _event("test", {})
        # Должен парситься как ISO
        datetime.fromisoformat(evt["timestamp"])

    def test_chat_conversation_updated_payload(self, monkeypatch):
        """Management update event should include action and conversation metadata."""
        sent = {}

        async def fake_send(member_user_ids, message, exclude_user_id=None):
            sent["member_user_ids"] = member_user_ids
            sent["message"] = message
            sent["exclude_user_id"] = exclude_user_id
            return len(member_user_ids)

        monkeypatch.setattr(
            "app.services.websocket_manager.ws_manager.send_to_conversation",
            fake_send,
        )

        run_async(
            broadcast_chat_conversation_updated(
                member_user_ids=[1, 2, 3],
                conversation_id=99,
                action="member_role_updated",
                actor_user_id=1,
                target_user_id=2,
                role="admin",
            )
        )

        assert sent["member_user_ids"] == [1, 2, 3]
        assert sent["exclude_user_id"] is None
        assert sent["message"]["type"] == "chat_conversation_updated"
        assert sent["message"]["data"] == {
            "conversation_id": 99,
            "action": "member_role_updated",
            "actor_user_id": 1,
            "target_user_id": 2,
            "role": "admin",
        }


# ============================================================================
# WebSocket API — Endpoint Tests
# ============================================================================


class TestWebSocketEndpoint:
    """Тесты WebSocket эндпоинта."""

    def test_ws_without_token(self, client):
        """WebSocket без токена → closed 4001."""
        with pytest.raises(Exception):
            with client.websocket_connect("/ws?token="):
                pass

    def test_ws_invalid_token(self, client):
        """WebSocket с невалидным токеном → closed 4001."""
        with pytest.raises(Exception):
            with client.websocket_connect("/ws?token=invalid_jwt_here"):
                pass

    def test_ws_valid_connection(self, client, admin_token):
        """WebSocket с валидным токеном — подключение и ping/pong."""
        with client.websocket_connect(f"/ws?token={admin_token}") as ws:
            ws.send_json({"type": "ping"})
            data = ws.receive_json()
            assert data["type"] == "pong"

    def test_ws_refresh_token_rejected(self, client, admin_user):
        """WebSocket не должен принимать refresh token."""
        from app.services.auth import create_refresh_token

        refresh_token = create_refresh_token(
            {"sub": admin_user.username, "user_id": admin_user.id}
        )

        with pytest.raises(Exception):
            with client.websocket_connect(f"/ws?token={refresh_token}"):
                pass


# ============================================================================
# Prometheus Metrics — Tests
# ============================================================================


class TestPrometheusMetrics:
    """Тесты Prometheus метрик."""

    def test_metrics_endpoint_exists(self, client):
        """Эндпоинт /metrics доступен."""
        response = client.get("/metrics")
        assert response.status_code == 200

    def test_metrics_format(self, client):
        """Метрики в формате Prometheus (text/plain)."""
        response = client.get("/metrics")
        content_type = response.headers.get("content-type", "")
        assert "text/plain" in content_type or "text/openmetrics" in content_type

    def test_metrics_contain_http_requests(self, client, client_with_auth):
        """Метрики содержат http_requests после запросов."""
        # Сделаем запрос, чтобы метрики появились
        client_with_auth.get("/api/sla?period=month")

        response = client.get("/metrics")
        body = response.text
        assert "http_request" in body or "http_requests" in body


# ============================================================================
# Health Check — WebSocket status
# ============================================================================


class TestHealthWebSocket:
    """WebSocket status в /health/detailed."""

    def test_health_detailed_has_websocket(self, client):
        """health/detailed содержит WebSocket статус."""
        response = client.get("/health/detailed")
        assert response.status_code == 200
        data = response.json()
        assert "websocket" in data
        assert "active_connections" in data["websocket"]
        assert "unique_users" in data["websocket"]
