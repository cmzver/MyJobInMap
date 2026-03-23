"""Regression tests for multi-tenant isolation between organizations."""

import io
import os

from fastapi.testclient import TestClient

from app.config import settings
from app.models import (AddressContactModel, AddressDocumentModel,
                        AddressHistoryEventType, AddressHistoryModel,
                        AddressModel, AddressSystemModel, OrganizationModel,
                        TaskModel, TaskPhotoModel, UserModel, UserRole)
from app.services.auth import get_password_hash


def _login_headers(client: TestClient, username: str, password: str) -> dict[str, str]:
    response = client.post(
        "/api/auth/login",
        data={"username": username, "password": password},
    )
    assert response.status_code == 200
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _create_org_user(
    db_session, *, username: str, password: str, role: str, organization_id: int
) -> UserModel:
    user = UserModel(
        username=username,
        password_hash=get_password_hash(password),
        full_name=username,
        role=role,
        is_active=True,
        organization_id=organization_id,
    )
    db_session.add(user)
    db_session.commit()
    db_session.refresh(user)
    return user


def test_org_admin_sees_only_own_tasks_and_users(client: TestClient, db_session):
    org1 = OrganizationModel(name="Org One", slug="org-one")
    org2 = OrganizationModel(name="Org Two", slug="org-two")
    db_session.add_all([org1, org2])
    db_session.commit()
    db_session.refresh(org1)
    db_session.refresh(org2)

    admin1 = _create_org_user(
        db_session,
        username="orgadmin1",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org1.id,
    )
    worker1 = _create_org_user(
        db_session,
        username="orgworker1",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org1.id,
    )
    _create_org_user(
        db_session,
        username="orgadmin2",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org2.id,
    )
    worker2 = _create_org_user(
        db_session,
        username="orgworker2",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org2.id,
    )

    task1 = TaskModel(
        title="Org1 Task",
        raw_address="Addr 1",
        status="NEW",
        priority="CURRENT",
        assigned_user_id=worker1.id,
        organization_id=org1.id,
    )
    task2 = TaskModel(
        title="Org2 Task",
        raw_address="Addr 2",
        status="NEW",
        priority="CURRENT",
        assigned_user_id=worker2.id,
        organization_id=org2.id,
    )
    db_session.add_all([task1, task2])
    db_session.commit()
    db_session.refresh(task1)
    db_session.refresh(task2)

    headers1 = _login_headers(client, admin1.username, "pass123")

    tasks_response = client.get("/api/tasks", headers=headers1)
    assert tasks_response.status_code == 200
    tasks = tasks_response.json()["items"]
    assert [task["id"] for task in tasks] == [task1.id]

    users_response = client.get("/api/admin/users", headers=headers1)
    assert users_response.status_code == 200
    usernames = {user["username"] for user in users_response.json()}
    assert usernames == {admin1.username, worker1.username}

    foreign_task_response = client.get(f"/api/tasks/{task2.id}", headers=headers1)
    assert foreign_task_response.status_code == 403


def test_org_admin_cannot_modify_or_assign_foreign_task(client: TestClient, db_session):
    org1 = OrganizationModel(name="Tenant A", slug="tenant-a")
    org2 = OrganizationModel(name="Tenant B", slug="tenant-b")
    db_session.add_all([org1, org2])
    db_session.commit()
    db_session.refresh(org1)
    db_session.refresh(org2)

    admin1 = _create_org_user(
        db_session,
        username="tenantadmin1",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org1.id,
    )
    _create_org_user(
        db_session,
        username="tenantworker1",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org1.id,
    )
    worker2 = _create_org_user(
        db_session,
        username="tenantworker2",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org2.id,
    )

    foreign_task = TaskModel(
        title="Foreign Task",
        raw_address="Foreign Addr",
        status="NEW",
        priority="CURRENT",
        assigned_user_id=worker2.id,
        organization_id=org2.id,
    )
    db_session.add(foreign_task)
    db_session.commit()
    db_session.refresh(foreign_task)

    headers1 = _login_headers(client, admin1.username, "pass123")

    update_response = client.patch(
        f"/api/admin/tasks/{foreign_task.id}",
        json={"title": "Hacked"},
        headers=headers1,
    )
    assert update_response.status_code == 404

    assign_response = client.patch(
        f"/api/tasks/{foreign_task.id}/assign",
        json={"assigned_user_id": worker2.id},
        headers=headers1,
    )
    assert assign_response.status_code == 403


def test_org_admin_cannot_access_foreign_address_or_user_stats(
    client: TestClient, db_session
):
    org1 = OrganizationModel(name="Addr Org 1", slug="addr-org-1")
    org2 = OrganizationModel(name="Addr Org 2", slug="addr-org-2")
    db_session.add_all([org1, org2])
    db_session.commit()
    db_session.refresh(org1)
    db_session.refresh(org2)

    admin1 = _create_org_user(
        db_session,
        username="addradmin1",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org1.id,
    )
    worker2 = _create_org_user(
        db_session,
        username="addrworker2",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org2.id,
    )

    foreign_address = AddressModel(
        address="Чужая улица, 5",
        city="Test",
        organization_id=org2.id,
    )
    db_session.add(foreign_address)
    db_session.commit()
    db_session.refresh(foreign_address)

    headers1 = _login_headers(client, admin1.username, "pass123")

    address_response = client.get(
        f"/api/addresses/{foreign_address.id}/full", headers=headers1
    )
    assert address_response.status_code == 403

    stats_response = client.get(
        f"/api/admin/users/{worker2.id}/stats", headers=headers1
    )
    assert stats_response.status_code == 403


def test_task_created_from_text_is_bound_to_admin_organization(
    client: TestClient, db_session
):
    org = OrganizationModel(name="Text Org", slug="text-org")
    db_session.add(org)
    db_session.commit()
    db_session.refresh(org)

    admin = _create_org_user(
        db_session,
        username="textadmin",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org.id,
    )

    headers = _login_headers(client, admin.username, "pass123")
    response = client.post(
        "/api/tasks/from-text",
        json={
            "text": "ул. Пушкина, дом 1\nНе работает домофон",
            "source": "web",
        },
        headers=headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["success"] is True

    task_id = data["task"]["id"]
    task = db_session.query(TaskModel).filter(TaskModel.id == task_id).first()
    assert task is not None
    assert task.organization_id == org.id


def test_address_created_by_org_admin_is_bound_and_visible_in_own_list(
    client: TestClient, db_session
):
    org = OrganizationModel(name="Address Tenant", slug="address-tenant")
    db_session.add(org)
    db_session.commit()
    db_session.refresh(org)

    admin = _create_org_user(
        db_session,
        username="addressadmin",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org.id,
    )

    headers = _login_headers(client, admin.username, "pass123")
    create_response = client.post(
        "/api/addresses",
        json={
            "address": "Tenant Visible St, 10",
            "city": "TenantCity",
            "street": "Tenant Visible St",
            "building": "10",
            "lat": 55.0,
            "lon": 37.0,
        },
        headers=headers,
    )

    assert create_response.status_code == 201
    address_id = create_response.json()["id"]

    address = (
        db_session.query(AddressModel).filter(AddressModel.id == address_id).first()
    )
    assert address is not None
    assert address.organization_id == org.id

    list_response = client.get(
        "/api/addresses?search=Tenant Visible St", headers=headers
    )
    assert list_response.status_code == 200
    items = list_response.json()["items"]
    assert [item["id"] for item in items] == [address_id]


def test_address_autocomplete_is_isolated_between_organizations(
    client: TestClient, db_session
):
    org1 = OrganizationModel(name="Autocomplete Org 1", slug="autocomplete-org-1")
    org2 = OrganizationModel(name="Autocomplete Org 2", slug="autocomplete-org-2")
    db_session.add_all([org1, org2])
    db_session.commit()
    db_session.refresh(org1)
    db_session.refresh(org2)

    admin1 = _create_org_user(
        db_session,
        username="autoadmin1",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org1.id,
    )
    admin2 = _create_org_user(
        db_session,
        username="autoadmin2",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org2.id,
    )

    own_address = AddressModel(
        address="Own Tenant Street, 1",
        city="OwnCity",
        street="Own Tenant Street",
        building="1",
        is_active=True,
        organization_id=org1.id,
    )
    foreign_address = AddressModel(
        address="Foreign Leak Street, 99",
        city="ForeignCity",
        street="Foreign Leak Street",
        building="99",
        is_active=True,
        organization_id=org2.id,
    )
    db_session.add_all([own_address, foreign_address])
    db_session.commit()

    headers1 = _login_headers(client, admin1.username, "pass123")
    headers2 = _login_headers(client, admin2.username, "pass123")

    cities_response = client.get(
        "/api/addresses/autocomplete/cities?q=Foreign", headers=headers1
    )
    assert cities_response.status_code == 200
    assert cities_response.json() == []

    full_response = client.get(
        "/api/addresses/autocomplete/full?q=Foreign Leak", headers=headers1
    )
    assert full_response.status_code == 200
    assert full_response.json() == []

    foreign_org_response = client.get(
        "/api/addresses/autocomplete/full?q=Foreign Leak", headers=headers2
    )
    assert foreign_org_response.status_code == 200
    assert [item["address"] for item in foreign_org_response.json()] == [
        "Foreign Leak Street, 99"
    ]


def test_org_admin_cannot_access_foreign_address_history_or_nested_mutations(
    client: TestClient, db_session
):
    org1 = OrganizationModel(name="Nested Org 1", slug="nested-org-1")
    org2 = OrganizationModel(name="Nested Org 2", slug="nested-org-2")
    db_session.add_all([org1, org2])
    db_session.commit()
    db_session.refresh(org1)
    db_session.refresh(org2)

    admin1 = _create_org_user(
        db_session,
        username="nestedadmin1",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org1.id,
    )
    admin2 = _create_org_user(
        db_session,
        username="nestedadmin2",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org2.id,
    )

    foreign_address = AddressModel(
        address="Nested Foreign Street, 7",
        city="NestedCity",
        organization_id=org2.id,
    )
    db_session.add(foreign_address)
    db_session.commit()
    db_session.refresh(foreign_address)

    foreign_system = AddressSystemModel(
        address_id=foreign_address.id,
        system_type="intercom",
        name="Foreign Intercom",
        status="active",
    )
    foreign_contact = AddressContactModel(
        address_id=foreign_address.id,
        contact_type="chairman",
        name="Foreign Contact",
        is_primary=True,
    )
    history_entry = AddressHistoryModel(
        address_id=foreign_address.id,
        event_type=AddressHistoryEventType.SYSTEM_ADDED.value,
        description="Foreign history entry",
        user_id=admin2.id,
    )

    uploads_dir = os.path.join(
        settings.BASE_DIR, "uploads", "address_documents", str(foreign_address.id)
    )
    os.makedirs(uploads_dir, exist_ok=True)
    file_name = "foreign-doc.txt"
    file_path = os.path.join(uploads_dir, file_name)
    with open(file_path, "w", encoding="utf-8") as file_handle:
        file_handle.write("tenant isolated document")

    foreign_document = AddressDocumentModel(
        address_id=foreign_address.id,
        name="Foreign Document",
        doc_type="other",
        file_path=f"/uploads/address_documents/{foreign_address.id}/{file_name}",
        file_size=24,
        mime_type="text/plain",
        created_by_id=admin2.id,
    )

    db_session.add_all(
        [foreign_system, foreign_contact, history_entry, foreign_document]
    )
    db_session.commit()
    db_session.refresh(foreign_system)
    db_session.refresh(foreign_document)

    headers1 = _login_headers(client, admin1.username, "pass123")

    history_response = client.get(
        f"/api/addresses/{foreign_address.id}/history", headers=headers1
    )
    assert history_response.status_code == 403

    update_system_response = client.patch(
        f"/api/addresses/{foreign_address.id}/systems/{foreign_system.id}",
        json={"name": "Hacked name"},
        headers=headers1,
    )
    assert update_system_response.status_code == 403

    delete_document_response = client.delete(
        f"/api/addresses/{foreign_address.id}/documents/{foreign_document.id}",
        headers=headers1,
    )
    assert delete_document_response.status_code == 403

    db_session.refresh(foreign_system)
    assert foreign_system.name == "Foreign Intercom"
    assert (
        db_session.query(AddressDocumentModel)
        .filter(AddressDocumentModel.id == foreign_document.id)
        .first()
        is not None
    )


def test_reports_exports_are_isolated_between_organizations(
    client: TestClient, db_session
):
    from openpyxl import load_workbook

    org1 = OrganizationModel(name="Reports Org 1", slug="reports-org-1")
    org2 = OrganizationModel(name="Reports Org 2", slug="reports-org-2")
    db_session.add_all([org1, org2])
    db_session.commit()
    db_session.refresh(org1)
    db_session.refresh(org2)

    admin1 = _create_org_user(
        db_session,
        username="reportsadmin1",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org1.id,
    )
    worker1 = _create_org_user(
        db_session,
        username="reportsworker1",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org1.id,
    )
    worker2 = _create_org_user(
        db_session,
        username="reportsworker2",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org2.id,
    )

    own_task = TaskModel(
        title="Org1 Visible Task",
        raw_address="Own report addr",
        status="DONE",
        priority="CURRENT",
        assigned_user_id=worker1.id,
        organization_id=org1.id,
    )
    foreign_task = TaskModel(
        title="Org2 Hidden Task",
        raw_address="Foreign report addr",
        status="DONE",
        priority="CURRENT",
        assigned_user_id=worker2.id,
        organization_id=org2.id,
    )
    db_session.add_all([own_task, foreign_task])
    db_session.commit()

    headers1 = _login_headers(client, admin1.username, "pass123")

    csv_response = client.get("/api/reports/export?period=all", headers=headers1)
    assert csv_response.status_code == 200
    csv_text = csv_response.text
    assert "Org1 Visible Task" in csv_text
    assert "Org2 Hidden Task" not in csv_text

    excel_response = client.get(
        "/api/reports/export/excel?period=all", headers=headers1
    )
    assert excel_response.status_code == 200
    workbook = load_workbook(io.BytesIO(excel_response.content))
    worksheet = workbook["Заявки"]
    titles = [
        worksheet.cell(row=row_idx, column=3).value
        for row_idx in range(2, worksheet.max_row + 1)
    ]
    assert "Org1 Visible Task" in titles
    assert "Org2 Hidden Task" not in titles


def test_reports_worker_filter_rejects_foreign_worker(client: TestClient, db_session):
    org1 = OrganizationModel(name="Worker Filter Org 1", slug="worker-filter-org-1")
    org2 = OrganizationModel(name="Worker Filter Org 2", slug="worker-filter-org-2")
    db_session.add_all([org1, org2])
    db_session.commit()
    db_session.refresh(org1)
    db_session.refresh(org2)

    admin1 = _create_org_user(
        db_session,
        username="workerfilteradmin1",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org1.id,
    )
    foreign_worker = _create_org_user(
        db_session,
        username="workerfilterforeign",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org2.id,
    )

    headers1 = _login_headers(client, admin1.username, "pass123")
    response = client.get(
        f"/api/reports?period=all&worker_id={foreign_worker.id}", headers=headers1
    )
    assert response.status_code == 404


def test_org_admin_cannot_delete_foreign_photo(client: TestClient, db_session):
    org1 = OrganizationModel(name="Photo Org 1", slug="photo-org-1")
    org2 = OrganizationModel(name="Photo Org 2", slug="photo-org-2")
    db_session.add_all([org1, org2])
    db_session.commit()
    db_session.refresh(org1)
    db_session.refresh(org2)

    admin1 = _create_org_user(
        db_session,
        username="photoadmin1",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org1.id,
    )
    admin2 = _create_org_user(
        db_session,
        username="photoadmin2",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org2.id,
    )

    foreign_task = TaskModel(
        title="Foreign Photo Task",
        raw_address="Photo addr",
        status="NEW",
        priority="CURRENT",
        organization_id=org2.id,
    )
    db_session.add(foreign_task)
    db_session.commit()
    db_session.refresh(foreign_task)

    photo_filename = "foreign-photo.jpg"
    photo_path = settings.PHOTOS_DIR / photo_filename
    photo_path.write_bytes(b"test-photo-bytes")

    foreign_photo = TaskPhotoModel(
        task_id=foreign_task.id,
        filename=photo_filename,
        original_name=photo_filename,
        file_size=16,
        mime_type="image/jpeg",
        photo_type="completion",
        uploaded_by_id=admin2.id,
    )
    db_session.add(foreign_photo)
    db_session.commit()
    db_session.refresh(foreign_photo)

    headers1 = _login_headers(client, admin1.username, "pass123")
    delete_response = client.delete(f"/api/photos/{foreign_photo.id}", headers=headers1)
    assert delete_response.status_code == 403
    assert photo_path.exists()
    assert (
        db_session.query(TaskPhotoModel)
        .filter(TaskPhotoModel.id == foreign_photo.id)
        .first()
        is not None
    )

    photo_path.unlink(missing_ok=True)


def test_v2_task_summary_is_isolated_between_organizations(
    client: TestClient, db_session
):
    org1 = OrganizationModel(name="V2 Summary Org 1", slug="v2-summary-org-1")
    org2 = OrganizationModel(name="V2 Summary Org 2", slug="v2-summary-org-2")
    db_session.add_all([org1, org2])
    db_session.commit()
    db_session.refresh(org1)
    db_session.refresh(org2)

    admin1 = _create_org_user(
        db_session,
        username="v2summaryadmin1",
        password="pass123",
        role=UserRole.ADMIN.value,
        organization_id=org1.id,
    )
    worker1 = _create_org_user(
        db_session,
        username="v2summaryworker1",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org1.id,
    )
    worker2 = _create_org_user(
        db_session,
        username="v2summaryworker2",
        password="pass123",
        role=UserRole.WORKER.value,
        organization_id=org2.id,
    )

    own_task = TaskModel(
        title="Visible v2 summary task",
        raw_address="Own v2 addr",
        status="NEW",
        priority="CURRENT",
        assigned_user_id=worker1.id,
        organization_id=org1.id,
    )
    foreign_task = TaskModel(
        title="Hidden v2 summary task",
        raw_address="Foreign v2 addr",
        status="DONE",
        priority="EMERGENCY",
        assigned_user_id=worker2.id,
        organization_id=org2.id,
    )
    db_session.add_all([own_task, foreign_task])
    db_session.commit()

    headers1 = _login_headers(client, admin1.username, "pass123")
    response = client.get("/api/v2/tasks/summary", headers=headers1)

    assert response.status_code == 200
    summary = response.json()["data"]
    assert summary["total"] == 1
    assert summary["by_status"] == {
        "NEW": 1,
        "IN_PROGRESS": 0,
        "DONE": 0,
        "CANCELLED": 0,
    }
    assert summary["by_priority"] == {
        "PLANNED": 0,
        "CURRENT": 1,
        "URGENT": 0,
        "EMERGENCY": 0,
    }
